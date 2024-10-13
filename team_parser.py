import openpyxl
import random

# "Team Intake Form Responses.xlsx"

# Download response sheet as a .xlsx file and replace the file path here
dataframe = openpyxl.load_workbook("/Users/alexandercolloredo-mansfeld/Assignments/Solution Challenge/Team Intake Form Responses.xlsx")

dataframe1 = dataframe.active

final_teams = []
full_teams = []
partial_teams = []
individuals = [["hello smello"], ["shmeep shmop"], ["bleep blop"]]

def parse_entry(entry):
    delim = ", "
    team = entry.split(delim)
    return team

def make_full_teams(partial_teams, individuals):
    full_teams = []
    split_partials = {
        2 : [],
        3 : [],
        4 : []
    }
    for pt in partial_teams:
        split_partials[len(pt)].append(pt)


    # Try to combine teams of two and three
    while len(split_partials[2]) > 0 and len(split_partials[3]) > 0:
        comb_team = []

        rand_pair = random.choice(split_partials[2])
        rand_triple = random.choice(split_partials[3])

        comb_team = rand_pair + rand_triple
        full_teams.append(comb_team)

        split_partials[2].remove(rand_pair)
        split_partials[3].remove(rand_triple)


    # Combine remaining teams with individuals
    for k, v in split_partials.items():
        while len(v) > 0:
            comb_team = []
            rand_partial = random.choice(v)

            for _ in range(5 - k):
                if len(individuals) > 0:
                    rand_indiv = random.choice(individuals)
                    comb_team += rand_indiv
                    individuals.remove(rand_indiv)

            comb_team += rand_partial
            full_teams.append(comb_team)

            v.remove(rand_partial)

    print(f"ALL INDIVIDUALS ASSIGNED A TEAM? {len(individuals) == 0}. {len(individuals)} LEFT OVER")
    # Group leftover individuals in groups of 5, and if there are less than 5, add to other random teams
    while len(individuals) >= 5:
        comb_team = []
        for i in range(5):
            rand_indiv = random.choice(individuals)
            comb_team += rand_indiv
            individuals.remove(rand_indiv)
        full_teams.append(comb_team)

    while len(individuals) > 0:
        rand_team_num = random.choice(range(0, len(full_teams)))
        rand_indiv = random.choice(individuals)
        full_teams[rand_team_num] += rand_indiv
        individuals.remove(rand_indiv)
    
    print(f"ALL INDIVIDUALS ASSIGNED A TEAM? {len(individuals) == 0}. {len(individuals)} LEFT OVER")

    return full_teams
                
total_members = 0
for row in range(1, dataframe1.max_row):
    for col_num, col in enumerate(dataframe1.iter_cols(2, dataframe1.max_column)):
        entry = col[row].value
        if entry == None:
            continue
        
        if col_num == 0:
            parsed_team = parse_entry(entry)
            full_teams.append(parsed_team)
            print(f"FULL TEAM WITH SIZE {len(parsed_team)}: {parsed_team}")
            total_members += len(parsed_team)

        if col_num == 1:
            parsed_team = parse_entry(entry)
            partial_teams.append(parsed_team)
            print(f"PARTIAL TEAM WITH SIZE {len(parsed_team)}: {parsed_team}")
            total_members += len(parsed_team)

        if col_num == 2:
            individuals.append([entry])
            print(f"INDIVIDUAL: {entry}")
            total_members += 1

print("\n\n\n")
print(f"TOTAL MEMBERS: {total_members}\n\n")

print(f"FULL TEAMS: {full_teams}\n")
print(f"PARTIAL TEAMS: {partial_teams}\n")
print(f"INDIVIDUALS: {individuals}\n")


generated_teams = make_full_teams(partial_teams, individuals)
print(f"GENERATED TEAMS: ")
for team in generated_teams:
    print(f"TEAM: {team}")

final_teams = full_teams + generated_teams
print("\n\n\nFINAL TEAMS:")
final_member_count = 0
for i, team in enumerate(final_teams):
    print(f"TEAM {i} WITH SIZE {len(team)}: {team}")
    final_member_count += len(team)

print(f"\n\n{final_member_count} ACCOUNTED FOR")